#!/usr/bin/env python3
"""
Migrate portfolio data from Excel to Google Sheets.

Run once after setting up your Google Sheet and service account:
    python migrate_to_sheets.py

Prerequisites:
    - GOOGLE_SHEET_ID env var set (or in .env)
    - credentials.json in the project folder (or GOOGLE_CREDENTIALS_JSON env var)
    - The Google Sheet must exist and be shared with your service account email
"""

import datetime
import os
import sys
from pathlib import Path

import yaml
from openpyxl import load_workbook

# Load config
ROOT = Path(__file__).resolve().parent
CFG = yaml.safe_load((ROOT / "config.yaml").read_text())
DATA_FILE = ROOT / CFG["data_file"]
BROKERS = CFG["brokers"]
BENCHMARKS = CFG["benchmarks"]

# Try loading .env if it exists
env_file = ROOT / ".env"
if env_file.exists():
    for line in env_file.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, val = line.split("=", 1)
            os.environ.setdefault(key.strip(), val.strip())

import gspread
from google.oauth2.service_account import Credentials

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


def migrate():
    if not DATA_FILE.exists():
        print(f"No Excel file found at {DATA_FILE}")
        sys.exit(1)

    sheet_id = os.environ.get("GOOGLE_SHEET_ID")
    if not sheet_id:
        print("Set GOOGLE_SHEET_ID env var first.")
        print("It's the ID from: https://docs.google.com/spreadsheets/d/<ID>/edit")
        sys.exit(1)

    # Auth
    creds_path = os.environ.get("GOOGLE_CREDENTIALS_JSON",
                                 str(ROOT / "credentials.json"))
    creds = Credentials.from_service_account_file(creds_path, scopes=SCOPES)
    client = gspread.authorize(creds)
    ss = client.open_by_key(sheet_id)

    wb = load_workbook(DATA_FILE, data_only=True)
    broker_names = [b["name"] for b in BROKERS]
    bench_labels = [b["label"] for b in BENCHMARKS]

    # ── Portfolio sheet ──
    print("Migrating Portfolio sheet...")
    ws_xl = wb["Portfolio"]
    header = ["Date"] + broker_names + ["Total"]

    try:
        ws_gs = ss.worksheet("Portfolio")
        ws_gs.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws_gs = ss.add_worksheet("Portfolio", rows=1, cols=len(header))

    rows = [header]
    for r in range(2, ws_xl.max_row + 1):
        d = ws_xl.cell(row=r, column=1).value
        if not d:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        row = [d.isoformat()]
        total = 0
        for c in range(2, 2 + len(broker_names)):
            v = ws_xl.cell(row=r, column=c).value
            val = float(v) if v is not None and not isinstance(v, str) else 0
            row.append(str(val))
            total += val
        row.append(str(total))
        rows.append(row)

    ws_gs.update(values=rows, range_name=f"A1:${chr(64+len(header))}{len(rows)}")
    print(f"  {len(rows)-1} rows written.")

    # ── Benchmarks sheet ──
    print("Migrating Benchmarks sheet...")
    ws_xl = wb["Benchmarks"]
    header = ["Date"] + bench_labels

    try:
        ws_gs = ss.worksheet("Benchmarks")
        ws_gs.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws_gs = ss.add_worksheet("Benchmarks", rows=1, cols=len(header))

    rows = [header]
    for r in range(2, ws_xl.max_row + 1):
        d = ws_xl.cell(row=r, column=1).value
        if not d:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        row = [d.isoformat()]
        for c in range(2, 2 + len(bench_labels)):
            v = ws_xl.cell(row=r, column=c).value
            row.append(str(float(v)) if v is not None and not isinstance(v, str) else "")
        rows.append(row)

    ws_gs.update(values=rows, range_name=f"A1:${chr(64+len(header))}{len(rows)}")
    print(f"  {len(rows)-1} rows written.")

    # ── CashFlows sheet ──
    if "CashFlows" in wb.sheetnames:
        print("Migrating CashFlows sheet...")
        ws_xl = wb["CashFlows"]
        header = ["Date", "Broker", "Amount", "Type"]

        try:
            ws_gs = ss.worksheet("CashFlows")
            ws_gs.clear()
        except gspread.exceptions.WorksheetNotFound:
            ws_gs = ss.add_worksheet("CashFlows", rows=1, cols=4)

        rows = [header]
        for r in range(2, ws_xl.max_row + 1):
            d = ws_xl.cell(row=r, column=1).value
            if not d:
                continue
            if isinstance(d, datetime.datetime):
                d = d.date()
            broker = ws_xl.cell(row=r, column=2).value or ""
            amount = ws_xl.cell(row=r, column=3).value
            flow_type = ws_xl.cell(row=r, column=4).value or ""
            rows.append([
                d.isoformat() if isinstance(d, datetime.date) else str(d),
                broker,
                str(float(amount)) if amount is not None else "0",
                flow_type,
            ])

        ws_gs.update(values=rows, range_name=f"A1:D{len(rows)}")
        print(f"  {len(rows)-1} rows written.")

    print("\nMigration complete!")


if __name__ == "__main__":
    migrate()
