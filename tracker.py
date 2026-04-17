#!/usr/bin/env python3
"""
Portfolio Tracker — daily performance vs benchmarks.

Usage:
    python tracker.py add            # interactively enter today's broker values
    python tracker.py fetch          # pull benchmark data from Yahoo Finance
    python tracker.py update         # add + fetch in one go
    python tracker.py dash           # regenerate the HTML dashboard
    python tracker.py demo           # seed 90 days of demo data (for testing)
    python tracker.py import-csv FILE.csv  # bulk-import portfolio values from CSV
    python tracker.py add-broker NAME [--currency EUR] [--type manual]
    python tracker.py remove-broker NAME
    python tracker.py list-brokers   # show configured brokers

The Excel file and HTML dashboard are configured in config.yaml.
"""

import argparse, datetime, sys, os, random
from pathlib import Path

import yaml
import yfinance as yf
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────
#  Config
# ──────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent
CFG = yaml.safe_load((ROOT / "config.yaml").read_text())

DATA_FILE = ROOT / CFG["data_file"]
DASH_FILE = ROOT / CFG["dashboard_file"]
BROKERS = CFG["brokers"]
BENCHMARKS = CFG["benchmarks"]

SHEET_PORTFOLIO = "Portfolio"
SHEET_BENCHMARKS = "Benchmarks"
SHEET_RETURNS = "Returns"

# ──────────────────────────────────────────────────
#  Excel helpers
# ──────────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2D3748")
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E8F0"),
)
PCT_FMT = '0.00%'
NUM_FMT = '#,##0.00'
DATE_FMT = 'YYYY-MM-DD'

def _style_header(ws, cols):
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(max_len + 3, 12)


def init_workbook():
    """Create a fresh data workbook with three sheets."""
    wb = Workbook()
    # Portfolio sheet: Date | Broker1 | Broker2 | … | Total
    ws = wb.active
    ws.title = SHEET_PORTFOLIO
    broker_names = [b["name"] for b in BROKERS]
    _style_header(ws, ["Date"] + broker_names + ["Total"])

    # Benchmarks sheet: Date | Bench1 | Bench2 | …
    ws2 = wb.create_sheet(SHEET_BENCHMARKS)
    bench_labels = [b["label"] for b in BENCHMARKS]
    _style_header(ws2, ["Date"] + bench_labels)

    # Returns sheet (cumulative % return): Date | Portfolio | Bench1 | …
    ws3 = wb.create_sheet(SHEET_RETURNS)
    _style_header(ws3, ["Date", "Portfolio"] + bench_labels)

    wb.save(DATA_FILE)
    print(f"✓ Created {DATA_FILE}")
    return wb


def load_or_create():
    if DATA_FILE.exists():
        return load_workbook(DATA_FILE)
    return init_workbook()


def _last_date(ws):
    """Return the last date in column A (or None)."""
    for row in range(ws.max_row, 1, -1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, datetime.datetime):
            return v.date()
        if isinstance(v, datetime.date):
            return v
    return None


# ──────────────────────────────────────────────────
#  ADD portfolio values
# ──────────────────────────────────────────────────
def cmd_add(date_str=None):
    wb = load_or_create()
    ws = wb[SHEET_PORTFOLIO]
    today = datetime.date.today() if not date_str else datetime.date.fromisoformat(date_str)

    last = _last_date(ws)
    if last and last >= today:
        print(f"⚠  Data for {today} already exists. Overwriting last row.")
        ws.delete_rows(ws.max_row)

    broker_names = [b["name"] for b in BROKERS]
    values = []
    print(f"\n📊  Enter portfolio values for {today}")
    print("─" * 40)
    for name in broker_names:
        while True:
            raw = input(f"  {name} value (€): ").strip().replace(",", ".")
            try:
                values.append(float(raw))
                break
            except ValueError:
                print("    Invalid number, try again.")

    total = sum(values)
    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=today).number_format = DATE_FMT
    for c, v in enumerate(values, 2):
        cell = ws.cell(row=row, column=c, value=v)
        cell.number_format = NUM_FMT
        cell.font = DATA_FONT
    total_col = len(broker_names) + 2
    # Use SUM formula
    first_col = get_column_letter(2)
    last_col = get_column_letter(total_col - 1)
    ws.cell(row=row, column=total_col, value=f"=SUM({first_col}{row}:{last_col}{row})")
    ws.cell(row=row, column=total_col).number_format = NUM_FMT
    ws.cell(row=row, column=total_col).font = Font(name="Arial", bold=True, size=10)

    _auto_width(ws)
    wb.save(DATA_FILE)
    print(f"✓ Saved portfolio total €{total:,.2f} for {today}")


def cmd_add_silent(date, values_dict):
    """Programmatic add (used by demo / import)."""
    wb = load_or_create()
    ws = wb[SHEET_PORTFOLIO]
    broker_names = [b["name"] for b in BROKERS]

    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=date).number_format = DATE_FMT
    total_val = 0
    for c, name in enumerate(broker_names, 2):
        v = values_dict.get(name, 0)
        total_val += v
        cell = ws.cell(row=row, column=c, value=v)
        cell.number_format = NUM_FMT
        cell.font = DATA_FONT

    total_col = len(broker_names) + 2
    first_col = get_column_letter(2)
    last_col = get_column_letter(total_col - 1)
    ws.cell(row=row, column=total_col, value=f"=SUM({first_col}{row}:{last_col}{row})")
    ws.cell(row=row, column=total_col).number_format = NUM_FMT
    ws.cell(row=row, column=total_col).font = Font(name="Arial", bold=True, size=10)

    _auto_width(ws)
    wb.save(DATA_FILE)
    return total_val


# ──────────────────────────────────────────────────
#  FETCH benchmark data
# ──────────────────────────────────────────────────
def cmd_fetch():
    wb = load_or_create()
    ws = wb[SHEET_BENCHMARKS]

    last = _last_date(ws)
    start = (last + datetime.timedelta(days=1)) if last else (datetime.date.today() - datetime.timedelta(days=365))
    end = datetime.date.today() + datetime.timedelta(days=1)

    if start >= end:
        print("Benchmarks already up to date.")
        return

    tickers = [b["ticker"] for b in BENCHMARKS]
    labels = [b["label"] for b in BENCHMARKS]
    print(f"📡  Fetching benchmarks from {start} to {datetime.date.today()} …")

    data = yf.download(tickers, start=str(start), end=str(end), progress=False, auto_adjust=True)

    if data.empty:
        print("⚠  No new benchmark data returned.")
        return

    close = data["Close"] if len(tickers) > 1 else data[["Close"]].rename(columns={"Close": tickers[0]})
    if isinstance(close.columns, pd.MultiIndex):
        close.columns = close.columns.droplevel(0)

    for dt_idx, row_data in close.iterrows():
        dt = dt_idx.date() if hasattr(dt_idx, "date") else dt_idx
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=dt).number_format = DATE_FMT
        for c, ticker in enumerate(tickers, 2):
            val = row_data.get(ticker, None)
            if pd.notna(val):
                cell = ws.cell(row=r, column=c, value=float(val))
                cell.number_format = NUM_FMT
                cell.font = DATA_FONT

    _auto_width(ws)
    wb.save(DATA_FILE)
    print(f"✓ Fetched {len(close)} days of benchmark data.")


# ──────────────────────────────────────────────────
#  COMPUTE cumulative returns sheet
# ──────────────────────────────────────────────────
def _compute_returns(wb):
    ws_p = wb[SHEET_PORTFOLIO]
    ws_b = wb[SHEET_BENCHMARKS]
    bench_labels = [b["label"] for b in BENCHMARKS]

    # Read portfolio dates+totals
    port = {}
    total_col = len(BROKERS) + 2
    for row in range(2, ws_p.max_row + 1):
        d = ws_p.cell(row=row, column=1).value
        v = ws_p.cell(row=row, column=total_col).value
        if d and v:
            if isinstance(d, datetime.datetime):
                d = d.date()
            port[d] = float(v) if not isinstance(v, str) else None

    # Read benchmark dates+values
    bench = {}
    for row in range(2, ws_b.max_row + 1):
        d = ws_b.cell(row=row, column=1).value
        if not d:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        vals = {}
        for c, label in enumerate(bench_labels, 2):
            v = ws_b.cell(row=row, column=c).value
            if v is not None and not isinstance(v, str):
                vals[label] = float(v)
        bench[d] = vals

    # For returns with formulas, we need matching dates
    # We'll compute returns in Python and write values (since these are derived analytics)
    all_dates = sorted(set(port.keys()) | set(bench.keys()))
    if not all_dates:
        return None

    # Build series
    port_series = {}
    bench_series = {l: {} for l in bench_labels}

    for d in all_dates:
        if d in port:
            port_series[d] = port[d]
        for l in bench_labels:
            if d in bench and l in bench[d]:
                bench_series[l][d] = bench[d][l]

    # Compute cumulative returns
    records = []
    port_dates = sorted(port_series.keys())
    base_port = port_series[port_dates[0]] if port_dates else None
    base_bench = {}
    for l in bench_labels:
        bd = sorted(bench_series[l].keys())
        if bd:
            # Find the base date that aligns with portfolio start
            for d in bd:
                if port_dates and d >= port_dates[0]:
                    base_bench[l] = bench_series[l][d]
                    break
            if l not in base_bench and bd:
                base_bench[l] = bench_series[l][bd[0]]

    for d in all_dates:
        rec = {"date": d}
        if d in port_series and base_port and base_port != 0:
            rec["Portfolio"] = (port_series[d] - base_port) / base_port
        for l in bench_labels:
            if d in bench_series[l] and l in base_bench and base_bench[l] != 0:
                rec[l] = (bench_series[l][d] - base_bench[l]) / base_bench[l]
        records.append(rec)

    # Write Returns sheet
    if SHEET_RETURNS in wb.sheetnames:
        del wb[SHEET_RETURNS]
    ws = wb.create_sheet(SHEET_RETURNS)
    cols = ["Date", "Portfolio"] + bench_labels
    _style_header(ws, cols)

    for rec in records:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=rec["date"]).number_format = DATE_FMT
        for c, col in enumerate(cols[1:], 2):
            v = rec.get(col)
            if v is not None:
                cell = ws.cell(row=row, column=c, value=v)
                cell.number_format = PCT_FMT
                cell.font = DATA_FONT

    _auto_width(ws)
    return records


# ──────────────────────────────────────────────────
#  DASHBOARD — generate interactive HTML
# ──────────────────────────────────────────────────
def cmd_dash():
    wb = load_or_create()
    records = _compute_returns(wb)
    wb.save(DATA_FILE)

    if not records:
        print("⚠  No data to plot. Run 'add' and 'fetch' first.")
        return

    bench_labels = [b["label"] for b in BENCHMARKS]
    broker_names = [b["name"] for b in BROKERS]

    # Read portfolio sheet for broker breakdown
    ws_p = wb[SHEET_PORTFOLIO]
    port_rows = []
    for row in range(2, ws_p.max_row + 1):
        d = ws_p.cell(row=row, column=1).value
        if not d:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        entry = {"date": str(d)}
        for c, name in enumerate(broker_names, 2):
            v = ws_p.cell(row=row, column=c).value
            entry[name] = float(v) if v and not isinstance(v, str) else 0
        total_col = len(broker_names) + 2
        v = ws_p.cell(row=row, column=total_col).value
        entry["Total"] = float(v) if v and not isinstance(v, str) else sum(entry.get(n, 0) for n in broker_names)
        port_rows.append(entry)

    # Prepare returns data for JS
    returns_js = []
    for rec in records:
        entry = {"date": str(rec["date"])}
        if "Portfolio" in rec:
            entry["Portfolio"] = round(rec["Portfolio"] * 100, 4)
        for l in bench_labels:
            if l in rec:
                entry[l] = round(rec[l] * 100, 4)
        returns_js.append(entry)

    import json
    returns_json = json.dumps(returns_js)
    portfolio_json = json.dumps(port_rows)
    benchmarks_json = json.dumps(bench_labels)
    brokers_json = json.dumps(broker_names)

    # Latest values for summary cards
    latest = port_rows[-1] if port_rows else {}
    first = port_rows[0] if port_rows else {}
    latest_total = latest.get("Total", 0)
    first_total = first.get("Total", 0)
    total_return = ((latest_total - first_total) / first_total * 100) if first_total else 0
    latest_ret = returns_js[-1] if returns_js else {}

    html = _generate_dashboard_html(
        returns_json=returns_json,
        portfolio_json=portfolio_json,
        benchmarks_json=benchmarks_json,
        brokers_json=brokers_json,
        latest_total=latest_total,
        total_return=total_return,
        latest_ret=latest_ret,
        bench_labels=bench_labels,
        broker_names=broker_names,
        latest=latest,
    )

    DASH_FILE.write_text(html, encoding="utf-8")
    print(f"✓ Dashboard written to {DASH_FILE}")


def _generate_dashboard_html(**ctx):
    bench_cards = ""
    for l in ctx["bench_labels"]:
        v = ctx["latest_ret"].get(l, 0)
        sign = "+" if v >= 0 else ""
        color = "#48BB78" if v >= 0 else "#F56565"
        bench_cards += f"""
        <div class="card">
            <div class="card-label">{l}</div>
            <div class="card-value" style="color:{color}">{sign}{v:.2f}%</div>
        </div>"""

    broker_cards = ""
    for name in ctx["broker_names"]:
        v = ctx["latest"].get(name, 0)
        broker_cards += f"""
        <div class="card">
            <div class="card-label">{name}</div>
            <div class="card-value">€{v:,.0f}</div>
        </div>"""

    port_sign = "+" if ctx["total_return"] >= 0 else ""
    port_color = "#48BB78" if ctx["total_return"] >= 0 else "#F56565"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Portfolio Tracker</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.min.js"></script>
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{
    font-family:'DM Sans',sans-serif;
    background:#0B0F19;
    color:#E2E8F0;
    min-height:100vh;
}}
.container {{ max-width:1280px; margin:0 auto; padding:32px 24px; }}

/* Header */
header {{
    display:flex; justify-content:space-between; align-items:flex-end;
    margin-bottom:40px; padding-bottom:24px;
    border-bottom:1px solid rgba(255,255,255,0.06);
}}
h1 {{
    font-size:28px; font-weight:700; letter-spacing:-0.5px;
    background:linear-gradient(135deg,#63B3ED,#B794F4);
    -webkit-background-clip:text; -webkit-text-fill-color:transparent;
}}
.subtitle {{ color:#718096; font-size:13px; margin-top:4px; }}
.header-right {{ text-align:right; }}
.total-value {{ font-size:36px; font-weight:700; letter-spacing:-1px;
    font-family:'JetBrains Mono',monospace; }}
.total-return {{ font-size:16px; font-weight:600;
    font-family:'JetBrains Mono',monospace; color:{port_color}; }}

/* Cards */
.cards {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(180px,1fr));
    gap:12px; margin-bottom:32px; }}
.card {{
    background:rgba(255,255,255,0.03); border:1px solid rgba(255,255,255,0.06);
    border-radius:12px; padding:16px 18px;
}}
.card-label {{ font-size:11px; text-transform:uppercase; letter-spacing:1px;
    color:#718096; margin-bottom:6px; }}
.card-value {{ font-size:20px; font-weight:600;
    font-family:'JetBrains Mono',monospace; }}

/* Sections */
.section {{ margin-bottom:40px; }}
.section-title {{
    font-size:14px; text-transform:uppercase; letter-spacing:1.5px;
    color:#718096; margin-bottom:16px; font-weight:600;
}}
.chart-box {{
    background:rgba(255,255,255,0.02); border:1px solid rgba(255,255,255,0.06);
    border-radius:16px; padding:24px; position:relative;
}}
canvas {{ width:100%!important; }}

/* Controls */
.controls {{
    display:flex; gap:8px; margin-bottom:16px; flex-wrap:wrap;
}}
.btn {{
    background:rgba(255,255,255,0.06); border:1px solid rgba(255,255,255,0.1);
    color:#CBD5E0; padding:6px 16px; border-radius:8px; font-size:12px;
    font-family:'DM Sans',sans-serif; cursor:pointer; transition:all .2s;
    font-weight:500;
}}
.btn:hover {{ background:rgba(255,255,255,0.12); color:#fff; }}
.btn.active {{ background:rgba(99,179,237,0.15); border-color:#63B3ED;
    color:#63B3ED; }}

/* Table */
.table-wrap {{ overflow-x:auto; border-radius:12px;
    border:1px solid rgba(255,255,255,0.06); }}
table {{ width:100%; border-collapse:collapse; font-size:13px; }}
th {{ background:rgba(255,255,255,0.04); padding:12px 16px; text-align:left;
    font-size:11px; text-transform:uppercase; letter-spacing:1px;
    color:#718096; font-weight:600; }}
td {{ padding:10px 16px; border-top:1px solid rgba(255,255,255,0.04);
    font-family:'JetBrains Mono',monospace; font-size:12px; }}
tr:hover td {{ background:rgba(255,255,255,0.02); }}
.pos {{ color:#48BB78; }} .neg {{ color:#F56565; }}

/* Footer */
footer {{ text-align:center; color:#4A5568; font-size:11px; margin-top:40px; padding-top:20px;
    border-top:1px solid rgba(255,255,255,0.04); }}
</style>
</head>
<body>
<div class="container">

<header>
  <div>
    <h1>Portfolio Tracker</h1>
    <div class="subtitle">Performance vs. Benchmarks</div>
  </div>
  <div class="header-right">
    <div class="total-value">€{ctx["latest_total"]:,.0f}</div>
    <div class="total-return">{port_sign}{ctx["total_return"]:.2f}% cumulative return</div>
  </div>
</header>

<!-- Summary cards -->
<div class="cards">
  {broker_cards}
  {bench_cards}
</div>

<!-- Cumulative Returns Chart -->
<div class="section">
  <div class="section-title">Cumulative Returns (%)</div>
  <div class="controls" id="period-controls">
    <button class="btn" data-period="30">1M</button>
    <button class="btn" data-period="90">3M</button>
    <button class="btn" data-period="180">6M</button>
    <button class="btn active" data-period="0">ALL</button>
  </div>
  <div class="chart-box">
    <canvas id="returnsChart" height="340"></canvas>
  </div>
</div>

<!-- Portfolio Value Chart -->
<div class="section">
  <div class="section-title">Portfolio Value (€)</div>
  <div class="chart-box">
    <canvas id="valueChart" height="280"></canvas>
  </div>
</div>

<!-- Broker Allocation Chart -->
<div class="section">
  <div class="section-title">Broker Allocation</div>
  <div class="chart-box" style="max-width:420px">
    <canvas id="allocationChart" height="280"></canvas>
  </div>
</div>

<!-- Data Table -->
<div class="section">
  <div class="section-title">Daily Returns Log</div>
  <div class="table-wrap">
    <table id="returnsTable">
      <thead><tr><th>Date</th><th>Portfolio</th></tr></thead>
      <tbody></tbody>
    </table>
  </div>
</div>

<footer>Updated {datetime.date.today().isoformat()} · Portfolio Tracker</footer>
</div>

<script>
const returnsData = {ctx["returns_json"]};
const portfolioData = {ctx["portfolio_json"]};
const benchLabels = {ctx["benchmarks_json"]};
const brokerNames = {ctx["brokers_json"]};

const COLORS = {{
    Portfolio: '#63B3ED',
    'S&P 500':  '#F6AD55',
    'NASDAQ':   '#FC8181',
    'MSCI World':'#B794F4',
}};
const FALLBACK = ['#68D391','#F6E05E','#FEB2B2','#C4B5FD','#76E4F7'];

function getColor(label, i) {{
    return COLORS[label] || FALLBACK[i % FALLBACK.length];
}}

// ── Returns chart ──────────────────────
function buildReturnsChart(days) {{
    const filtered = days > 0 ? returnsData.slice(-days) : returnsData;
    const labels = filtered.map(r => r.date);
    const datasets = [];
    // Portfolio
    datasets.push({{
        label: 'Portfolio',
        data: filtered.map(r => r.Portfolio ?? null),
        borderColor: getColor('Portfolio', 0),
        backgroundColor: 'transparent',
        borderWidth: 2.5, pointRadius: 0, tension: 0.3,
    }});
    benchLabels.forEach((l, i) => {{
        datasets.push({{
            label: l,
            data: filtered.map(r => r[l] ?? null),
            borderColor: getColor(l, i+1),
            backgroundColor: 'transparent',
            borderWidth: 1.8, pointRadius: 0, tension: 0.3,
            borderDash: [6, 3],
        }});
    }});

    if (window._returnsChart) window._returnsChart.destroy();
    window._returnsChart = new Chart(document.getElementById('returnsChart'), {{
        type: 'line',
        data: {{ labels, datasets }},
        options: {{
            responsive: true,
            interaction: {{ intersect: false, mode: 'index' }},
            plugins: {{
                legend: {{ labels: {{ color:'#A0AEC0', font:{{ family:'DM Sans', size:12 }} }} }},
                tooltip: {{
                    backgroundColor:'rgba(26,32,44,0.95)', titleColor:'#E2E8F0',
                    bodyColor:'#CBD5E0', borderColor:'rgba(255,255,255,0.1)', borderWidth:1,
                    padding:12, titleFont:{{ family:'DM Sans' }}, bodyFont:{{ family:'JetBrains Mono', size:12 }},
                    callbacks: {{ label: ctx => ctx.dataset.label + ': ' + (ctx.parsed.y != null ? ctx.parsed.y.toFixed(2)+'%' : '—') }}
                }}
            }},
            scales: {{
                x: {{ ticks:{{ color:'#4A5568', font:{{ size:10 }}, maxTicksLimit:10 }},
                       grid:{{ color:'rgba(255,255,255,0.03)' }} }},
                y: {{ ticks:{{ color:'#4A5568', font:{{ family:'JetBrains Mono', size:11 }},
                             callback: v => v.toFixed(1)+'%' }},
                       grid:{{ color:'rgba(255,255,255,0.04)' }} }}
            }}
        }}
    }});
}}
buildReturnsChart(0);

document.querySelectorAll('#period-controls .btn').forEach(btn => {{
    btn.addEventListener('click', () => {{
        document.querySelectorAll('#period-controls .btn').forEach(b => b.classList.remove('active'));
        btn.classList.add('active');
        buildReturnsChart(+btn.dataset.period);
    }});
}});

// ── Portfolio value chart ──────────────
(() => {{
    const labels = portfolioData.map(r => r.date);
    const datasets = brokerNames.map((name, i) => ({{
        label: name,
        data: portfolioData.map(r => r[name] || 0),
        backgroundColor: getColor(name, i+2) + '99',
        borderColor: getColor(name, i+2),
        borderWidth: 1, fill: true, pointRadius: 0, tension: 0.3,
    }}));
    new Chart(document.getElementById('valueChart'), {{
        type: 'line',
        data: {{ labels, datasets }},
        options: {{
            responsive: true, interaction: {{ intersect:false, mode:'index' }},
            plugins: {{
                legend: {{ labels:{{ color:'#A0AEC0', font:{{ family:'DM Sans', size:12 }} }} }},
                tooltip: {{
                    backgroundColor:'rgba(26,32,44,0.95)', titleColor:'#E2E8F0',
                    bodyColor:'#CBD5E0', borderColor:'rgba(255,255,255,0.1)', borderWidth:1,
                    padding:12, bodyFont:{{ family:'JetBrains Mono', size:12 }},
                    callbacks: {{ label: ctx => ctx.dataset.label + ': €' + (ctx.parsed.y||0).toLocaleString() }}
                }}
            }},
            scales: {{
                x: {{ ticks:{{ color:'#4A5568', font:{{ size:10 }}, maxTicksLimit:10 }},
                       grid:{{ color:'rgba(255,255,255,0.03)' }} }},
                y: {{ stacked:false,
                       ticks:{{ color:'#4A5568', font:{{ family:'JetBrains Mono', size:11 }},
                               callback: v => '€'+v.toLocaleString() }},
                       grid:{{ color:'rgba(255,255,255,0.04)' }} }}
            }}
        }}
    }});
}})();

// ── Allocation doughnut ────────────────
(() => {{
    const latest = portfolioData[portfolioData.length - 1] || {{}};
    const vals = brokerNames.map(n => latest[n] || 0);
    new Chart(document.getElementById('allocationChart'), {{
        type: 'doughnut',
        data: {{
            labels: brokerNames,
            datasets: [{{ data: vals,
                backgroundColor: brokerNames.map((n,i) => getColor(n,i+2)),
                borderColor: '#0B0F19', borderWidth: 3 }}]
        }},
        options: {{
            responsive: true, cutout: '65%',
            plugins: {{
                legend: {{ position:'right', labels:{{ color:'#A0AEC0', font:{{ family:'DM Sans', size:12 }},
                    padding:16 }} }},
                tooltip: {{
                    backgroundColor:'rgba(26,32,44,0.95)', titleColor:'#E2E8F0',
                    bodyColor:'#CBD5E0', bodyFont:{{ family:'JetBrains Mono', size:12 }},
                    callbacks: {{ label: ctx => ctx.label + ': €' + ctx.parsed.toLocaleString() }}
                }}
            }}
        }}
    }});
}})();

// ── Data table ─────────────────────────
(() => {{
    const table = document.getElementById('returnsTable');
    const thead = table.querySelector('thead tr');
    benchLabels.forEach(l => {{
        const th = document.createElement('th');
        th.textContent = l;
        thead.appendChild(th);
    }});
    const tbody = table.querySelector('tbody');
    [...returnsData].reverse().slice(0, 60).forEach(r => {{
        const tr = document.createElement('tr');
        tr.innerHTML = `<td>${{r.date}}</td>`;
        ['Portfolio', ...benchLabels].forEach(key => {{
            const v = r[key];
            const cls = v > 0 ? 'pos' : v < 0 ? 'neg' : '';
            const txt = v != null ? (v > 0 ? '+' : '') + v.toFixed(2) + '%' : '—';
            tr.innerHTML += `<td class="${{cls}}">${{txt}}</td>`;
        }});
        tbody.appendChild(tr);
    }});
}})();
</script>
</body>
</html>"""


# ──────────────────────────────────────────────────
#  DEMO data generator
# ──────────────────────────────────────────────────
def cmd_demo():
    if DATA_FILE.exists():
        DATA_FILE.unlink()
    print("🎲  Generating 90 days of demo data …")

    base_values = {"DEGIRO": 15000, "XTB": 8000, "Renta4": 5000}
    today = datetime.date.today()
    start = today - datetime.timedelta(days=90)

    d = start
    while d <= today:
        if d.weekday() < 5:  # business days
            vals = {}
            for name, base in base_values.items():
                drift = random.gauss(0.0004, 0.012)
                base_values[name] = base * (1 + drift)
                vals[name] = round(base_values[name], 2)
            cmd_add_silent(d, vals)
        d += datetime.timedelta(days=1)

    cmd_fetch()
    cmd_dash()
    print("✓ Demo complete — open dashboard.html in your browser!")


# ──────────────────────────────────────────────────
#  IMPORT from CSV
# ──────────────────────────────────────────────────
def cmd_import_csv(filepath):
    """Import portfolio values from a CSV with columns: date, broker_name, value."""
    df = pd.read_csv(filepath, parse_dates=["date"])
    broker_names = [b["name"] for b in BROKERS]
    for dt, group in df.groupby("date"):
        vals = {}
        for _, row in group.iterrows():
            if row["broker_name"] in broker_names:
                vals[row["broker_name"]] = float(row["value"])
        cmd_add_silent(dt.date() if hasattr(dt, "date") else dt, vals)
    print(f"✓ Imported {len(df)} rows from {filepath}")


# ──────────────────────────────────────────────────
#  BROKER management (CLI wrappers)
# ──────────────────────────────────────────────────
def _cmd_add_broker(name, currency, broker_type):
    from viewer.data.excel_io import add_broker, _reload_config
    err = add_broker(name, currency, broker_type)
    if err:
        return err
    _reload_config()
    global BROKERS
    BROKERS = CFG["brokers"]
    print(f"✓ Added broker '{name}' ({currency}, {broker_type})")
    return None


def _cmd_remove_broker(name):
    from viewer.data.excel_io import remove_broker, _reload_config
    err = remove_broker(name)
    if err:
        return err
    _reload_config()
    global BROKERS
    BROKERS = CFG["brokers"]
    print(f"✓ Removed broker '{name}'")
    return None


def _cmd_list_brokers():
    from viewer.data.excel_io import list_brokers
    brokers = list_brokers()
    print(f"\n📋  Configured brokers ({len(brokers)}):")
    print("─" * 40)
    for b in brokers:
        print(f"  {b['name']:15s}  {b['currency']:5s}  ({b['type']})")
    print()


# ──────────────────────────────────────────────────
#  CLI
# ──────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Portfolio Tracker")
    sub = parser.add_subparsers(dest="command")

    sub.add_parser("add", help="Enter today's portfolio values")
    sub.add_parser("fetch", help="Fetch benchmark data")
    sub.add_parser("update", help="add + fetch + dash")
    sub.add_parser("dash", help="Regenerate HTML dashboard")
    sub.add_parser("demo", help="Generate demo data for testing")

    p_imp = sub.add_parser("import-csv", help="Bulk import from CSV")
    p_imp.add_argument("file", help="CSV file path")

    p_add_b = sub.add_parser("add-broker", help="Add a new broker")
    p_add_b.add_argument("name", help="Broker display name")
    p_add_b.add_argument("--currency", default="EUR", help="Currency (default: EUR)")
    p_add_b.add_argument("--type", default="manual", dest="broker_type",
                         help="Type: manual or selenium (default: manual)")

    p_rm_b = sub.add_parser("remove-broker", help="Remove a broker")
    p_rm_b.add_argument("name", help="Broker name to remove")

    sub.add_parser("list-brokers", help="List configured brokers")

    args = parser.parse_args()

    if args.command == "add":
        cmd_add()
    elif args.command == "fetch":
        cmd_fetch()
    elif args.command == "update":
        cmd_add()
        cmd_fetch()
        cmd_dash()
    elif args.command == "dash":
        cmd_dash()
    elif args.command == "demo":
        cmd_demo()
    elif args.command == "import-csv":
        cmd_import_csv(args.file)
    elif args.command == "add-broker":
        err = _cmd_add_broker(args.name, args.currency, args.broker_type)
        if err:
            print(f"⚠  {err}")
            sys.exit(1)
    elif args.command == "remove-broker":
        err = _cmd_remove_broker(args.name)
        if err:
            print(f"⚠  {err}")
            sys.exit(1)
    elif args.command == "list-brokers":
        _cmd_list_brokers()
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
