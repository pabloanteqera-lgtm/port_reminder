"""Excel I/O operations — save, fetch, ensure workbook exists."""

import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent.parent
CFG = yaml.safe_load((ROOT / "config.yaml").read_text())
DATA_FILE = ROOT / CFG["data_file"]
BROKERS = CFG["brokers"]
BENCHMARKS = CFG["benchmarks"]

SHEET_PORTFOLIO = "Portfolio"
SHEET_BENCHMARKS = "Benchmarks"
SHEET_RETURNS = "Returns"
SHEET_CASHFLOWS = "CashFlows"
NUM_FMT = '#,##0.00'
DATE_FMT = 'YYYY-MM-DD'


def _last_date(ws):
    for row in range(ws.max_row, 1, -1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, datetime.datetime):
            return v.date()
        if isinstance(v, datetime.date):
            return v
    return None


def _ensure_workbook():
    """Create the Excel file if it doesn't exist."""
    if DATA_FILE.exists():
        # Ensure CashFlows sheet exists in existing workbook
        wb = load_workbook(DATA_FILE)
        if SHEET_CASHFLOWS not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_CASHFLOWS)
            for c, h in enumerate(["Date", "Broker", "Amount", "Type"], 1):
                ws.cell(row=1, column=c, value=h)
            wb.save(DATA_FILE)
        return

    wb = Workbook()
    broker_names = [b["name"] for b in BROKERS]
    bench_labels = [b["label"] for b in BENCHMARKS]

    ws = wb.active
    ws.title = SHEET_PORTFOLIO
    for c, h in enumerate(["Date"] + broker_names + ["Total"], 1):
        ws.cell(row=1, column=c, value=h)

    ws2 = wb.create_sheet(SHEET_BENCHMARKS)
    for c, h in enumerate(["Date"] + bench_labels, 1):
        ws2.cell(row=1, column=c, value=h)

    ws3 = wb.create_sheet(SHEET_RETURNS)
    for c, h in enumerate(["Date", "Portfolio"] + bench_labels, 1):
        ws3.cell(row=1, column=c, value=h)

    ws4 = wb.create_sheet(SHEET_CASHFLOWS)
    for c, h in enumerate(["Date", "Broker", "Amount", "Type"], 1):
        ws4.cell(row=1, column=c, value=h)

    wb.save(DATA_FILE)


def fetch_benchmarks():
    """Fetch benchmark data from Yahoo Finance and save to Excel."""
    _ensure_workbook()
    wb = load_workbook(DATA_FILE)
    ws = wb[SHEET_BENCHMARKS]

    last = _last_date(ws)
    start = (last + datetime.timedelta(days=1)) if last else (
        datetime.date.today() - datetime.timedelta(days=365))
    end = datetime.date.today() + datetime.timedelta(days=1)

    if start >= end:
        return "Benchmarks up to date."

    import pandas as pd
    import yfinance as yf

    tickers = [b["ticker"] for b in BENCHMARKS]
    data = yf.download(tickers, start=str(start), end=str(end),
                       progress=False, auto_adjust=True)

    if data.empty:
        return "No new benchmark data."

    close = (data["Close"] if len(tickers) > 1
             else data[["Close"]].rename(columns={"Close": tickers[0]}))
    if isinstance(close.columns, pd.MultiIndex):
        close.columns = close.columns.droplevel(0)

    # Resample to weekly (Friday close) to match existing weekly data
    close = close.resample("W-FRI").last().dropna(how="all")

    count = 0
    for dt_idx, row_data in close.iterrows():
        dt = dt_idx.date() if hasattr(dt_idx, "date") else dt_idx
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=dt).number_format = DATE_FMT
        for c, ticker in enumerate(tickers, 2):
            val = row_data.get(ticker, None)
            if pd.notna(val):
                ws.cell(row=r, column=c, value=float(val)).number_format = NUM_FMT
        count += 1

    wb.save(DATA_FILE)
    return f"Fetched {count} days of benchmark data."


def save_portfolio_values(date, values_dict):
    """Save broker values for a given date to Excel."""
    _ensure_workbook()
    wb = load_workbook(DATA_FILE)
    ws = wb[SHEET_PORTFOLIO]
    broker_names = [b["name"] for b in BROKERS]

    last = _last_date(ws)
    if last and last >= date:
        for row in range(ws.max_row, 1, -1):
            d = ws.cell(row=row, column=1).value
            if isinstance(d, datetime.datetime):
                d = d.date()
            if d == date:
                ws.delete_rows(row)
                break

    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=date).number_format = DATE_FMT
    for c, name in enumerate(broker_names, 2):
        v = values_dict.get(name, 0)
        ws.cell(row=row, column=c, value=v).number_format = NUM_FMT

    total_col = len(broker_names) + 2
    first_col = get_column_letter(2)
    last_col = get_column_letter(total_col - 1)
    ws.cell(row=row, column=total_col,
            value=f"=SUM({first_col}{row}:{last_col}{row})")
    ws.cell(row=row, column=total_col).number_format = NUM_FMT

    wb.save(DATA_FILE)
    return sum(values_dict.values())


def save_cashflow(date, broker, amount, flow_type):
    """Save a deposit or withdrawal to the CashFlows sheet."""
    _ensure_workbook()
    wb = load_workbook(DATA_FILE)
    ws = wb[SHEET_CASHFLOWS]

    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=date).number_format = DATE_FMT
    ws.cell(row=row, column=2, value=broker)
    ws.cell(row=row, column=3, value=amount).number_format = NUM_FMT
    ws.cell(row=row, column=4, value=flow_type)

    wb.save(DATA_FILE)


def has_today_values():
    """Check if portfolio values for today already exist."""
    _ensure_workbook()
    wb = load_workbook(DATA_FILE, data_only=True)
    ws = wb[SHEET_PORTFOLIO]
    last = _last_date(ws)
    return last == datetime.date.today() if last else False


# ──────────────────────────────────────────────────
#  Broker management
# ──────────────────────────────────────────────────

def _reload_config():
    """Re-read config.yaml and update module-level BROKERS/BENCHMARKS."""
    global CFG, BROKERS, BENCHMARKS
    CFG = yaml.safe_load((ROOT / "config.yaml").read_text())
    BROKERS = CFG["brokers"]
    BENCHMARKS = CFG["benchmarks"]


def _save_config():
    """Write the current CFG back to config.yaml."""
    (ROOT / "config.yaml").write_text(yaml.dump(CFG, default_flow_style=False,
                                                 sort_keys=False,
                                                 allow_unicode=True))


def add_broker(name, currency="EUR", broker_type="manual"):
    """Add a new broker to config.yaml and insert a column in the Excel workbook."""
    _reload_config()

    # Check for duplicate name (case-insensitive)
    existing = [b["name"].lower() for b in BROKERS]
    if name.lower() in existing:
        return f"Broker '{name}' already exists."

    # 1. Update config
    CFG["brokers"].append({"name": name, "type": broker_type, "currency": currency})
    _save_config()
    _reload_config()

    # 2. Migrate Excel workbook — insert column before "Total"
    if DATA_FILE.exists():
        wb = load_workbook(DATA_FILE)
        ws = wb[SHEET_PORTFOLIO]

        # Total is the last header column
        total_col = ws.max_column  # current Total column
        # Insert a new column before Total
        ws.insert_cols(total_col)
        ws.cell(row=1, column=total_col, value=name)

        # Fix SUM formulas in the Total column (now shifted right by 1)
        new_total_col = total_col + 1
        first_col = get_column_letter(2)
        last_col = get_column_letter(new_total_col - 1)
        for row in range(2, ws.max_row + 1):
            # Set the new broker column to 0 for existing rows
            ws.cell(row=row, column=total_col, value=0).number_format = NUM_FMT
            # Rewrite the SUM formula
            ws.cell(row=row, column=new_total_col,
                    value=f"=SUM({first_col}{row}:{last_col}{row})")
            ws.cell(row=row, column=new_total_col).number_format = NUM_FMT

        wb.save(DATA_FILE)

    return None  # success


def remove_broker(name):
    """Remove a broker from config.yaml and delete its column from the Excel workbook."""
    _reload_config()

    broker_names = [b["name"] for b in BROKERS]
    if name not in broker_names:
        return f"Broker '{name}' not found."

    if len(broker_names) <= 1:
        return "Cannot remove the last broker."

    # Find the column index (1-based; Date=1, first broker=2, …)
    col_idx = broker_names.index(name) + 2  # +1 for 0-index, +1 for Date column

    # 1. Remove from config
    CFG["brokers"] = [b for b in CFG["brokers"] if b["name"] != name]
    _save_config()
    _reload_config()

    # 2. Migrate Excel workbook — delete the column and fix SUM formulas
    if DATA_FILE.exists():
        wb = load_workbook(DATA_FILE)
        ws = wb[SHEET_PORTFOLIO]

        ws.delete_cols(col_idx)

        # Rewrite SUM formulas in the (now shifted) Total column
        new_broker_count = len(CFG["brokers"])
        new_total_col = new_broker_count + 2
        first_col = get_column_letter(2)
        last_col = get_column_letter(new_total_col - 1)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=new_total_col,
                    value=f"=SUM({first_col}{row}:{last_col}{row})")
            ws.cell(row=row, column=new_total_col).number_format = NUM_FMT

        wb.save(DATA_FILE)

    return None  # success


def list_brokers():
    """Return the current broker list from config."""
    _reload_config()
    return BROKERS
