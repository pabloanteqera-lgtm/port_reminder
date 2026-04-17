"""Load data from Excel into Python dicts."""

import datetime
from openpyxl import load_workbook

from .excel_io import (DATA_FILE, BROKERS, BENCHMARKS, _ensure_workbook,
                       SHEET_PORTFOLIO, SHEET_BENCHMARKS, SHEET_CASHFLOWS)


def load_data():
    """Read Excel and return raw per-broker values and benchmark raw values."""
    _ensure_workbook()
    wb = load_workbook(DATA_FILE, data_only=True)
    bench_labels = [b["label"] for b in BENCHMARKS]
    broker_names = [b["name"] for b in BROKERS]

    ws_p = wb[SHEET_PORTFOLIO]
    broker_raw = {name: {} for name in broker_names}
    for row in range(2, ws_p.max_row + 1):
        d = ws_p.cell(row=row, column=1).value
        if not d:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        for i, name in enumerate(broker_names):
            bv = ws_p.cell(row=row, column=i + 2).value
            if bv is not None and not isinstance(bv, str):
                broker_raw[name][d] = float(bv)

    ws_b = wb[SHEET_BENCHMARKS]
    bench_raw = {l: {} for l in bench_labels}
    for row in range(2, ws_b.max_row + 1):
        d = ws_b.cell(row=row, column=1).value
        if not d:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()
        for c, label in enumerate(bench_labels, 2):
            v = ws_b.cell(row=row, column=c).value
            if v is not None and not isinstance(v, str):
                bench_raw[label][d] = float(v)

    # Forward-fill benchmarks to cover portfolio dates (weekends, holidays)
    all_portfolio_dates = sorted(
        {d for br in broker_raw.values() for d in br})
    for label, raw in bench_raw.items():
        if not raw:
            continue
        bench_dates = sorted(raw.keys())
        last_val = None
        bi = 0
        for d in all_portfolio_dates:
            while bi < len(bench_dates) and bench_dates[bi] <= d:
                last_val = raw[bench_dates[bi]]
                bi += 1
            if d not in raw and last_val is not None:
                raw[d] = last_val

    return broker_raw, bench_raw


def load_cashflows():
    """Load cash flows from Excel. Returns {broker: [(date, signed_amount), ...]}."""
    _ensure_workbook()
    wb = load_workbook(DATA_FILE, data_only=True)
    if SHEET_CASHFLOWS not in wb.sheetnames:
        return {}

    ws = wb[SHEET_CASHFLOWS]
    flows = {}
    for row in range(2, ws.max_row + 1):
        d = ws.cell(row=row, column=1).value
        broker = ws.cell(row=row, column=2).value
        amount = ws.cell(row=row, column=3).value
        flow_type = ws.cell(row=row, column=4).value

        if not d or not broker or amount is None:
            continue
        if isinstance(d, datetime.datetime):
            d = d.date()

        signed = float(amount) if flow_type == "deposit" else -float(amount)
        flows.setdefault(broker, []).append((d, signed))

    return flows
